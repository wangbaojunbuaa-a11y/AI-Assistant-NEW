from openpyxl import load_workbook
from .base import ExcelHandler

class PyxlDriver(ExcelHandler):
    """使用 openpyxl 处理后台 Excel 文件"""

    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)

    def get_value(self, sheet_name, cell_address):
        sheet = self.wb[sheet_name]
        return sheet[cell_address].value

    def set_value(self, sheet_name, cell_address, value):
        sheet = self.wb[sheet_name]
        sheet[cell_address].value = value

    def get_cell_by_index(self, sheet_name, row, col):
        sheet = self.wb[sheet_name]
        return sheet.cell(row=row, column=col).value

    def set_cell_by_index(self, sheet_name, row, col, value):
        sheet = self.wb[sheet_name]
        sheet.cell(row=row, column=col).value = value

    def get_sheet_count(self):
        return len(self.wb.sheetnames)

    def get_sheet_by_index(self, index):
        return self.wb.sheetnames[index]

    def capture_range_image(self, sheet_name, range_str, save_path):
        # openpyxl 不支持截图
        raise NotImplementedError("PyxlDriver 不支持截图功能，请使用 ComDriver")

    def copy_sheet_from_external(self, source_path, source_sheet_name, target_sheet_name, after_sheet_index=None):
        raise NotImplementedError("PyxlDriver 不支持跨文件复制工作表逻辑，请使用 ComDriver")

    def merge_cells(self, sheet_name, range_str):
        sheet = self.wb[sheet_name]
        sheet.merge_cells(range_str)

    def get_sheet_names(self):
        return self.wb.sheetnames

    def save(self, path=None):
        self.wb.save(path or self.path)

    def close(self):
        self.wb.close()
